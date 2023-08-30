from constants import *
import custom_logger
import openpyxl
import yaml
from openpyxl.styles import PatternFill, Font, numbers, Protection, Alignment
from openpyxl.formatting.rule import FormulaRule


class Forecaster():

    def __init__(self):
        with open('config.yaml', 'r') as f:
            self.config = yaml.safe_load(f)
        self.logger = custom_logger.get_logger(__name__)
        self.excel_num_columns_from_header = 3
        self.excel_work_item_section_start_row = 9
        self.excel_num_rows_between_features = 3
        self.iteration_capacity_row = 3

    def create_forecast_as_excel(self, work_item_detail_list, iteration_list, output_path):
        filtered_item_list = [x for x in work_item_detail_list if
                              x.get(WorkItemFields.WorkItemType) in [WorkItemTypes.Epic,
                                                                     WorkItemTypes.Feature]]

        num_epics = len([x for x in work_item_detail_list if
                         x.get(WorkItemFields.WorkItemType) == WorkItemTypes.Epic])
        num_features = len([x for x in work_item_detail_list if
                            x.get(WorkItemFields.WorkItemType) == WorkItemTypes.Feature])

        self.totals_row = self.__calculate_totals_row__(num_epics, num_features)

        self.logger.debug('start')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.protection.sheet = True

        ws.freeze_panes = 'D8'

        self.__write_iterations__(iteration_list, ws)
        self.__write_items__(filtered_item_list, iteration_list, ws)

        wb.save(output_path)
        self.logger.debug('end')

    def __write_iterations__(self, iteration_list, ws):
        self.logger.debug('start')
        ws['A1'] = 'Sprint'
        ws['A2'] = 'Dates'
        ws['A3'] = 'Capacity'
        ws['A4'] = 'Total Allocated'
        ws['A5'] = 'Overallocated?'

        format = self.config.get('Forecaster').get('VisualFormats').get('Iteration').get('Title')
        pattern_fill = PatternFill(**format.get('PatternFill'))
        font = Font(**format.get('Font'))

        for i in range(0, len(iteration_list)):
            cell_letter = self.__number_to_excel_column_letter__(i + self.excel_num_columns_from_header)

            cell = f'{cell_letter}1'
            ws[cell] = iteration_list[i].get(IterationFields.Name)
            ws[cell].fill = pattern_fill
            ws[cell].font = font

            cell = f'{cell_letter}2'
            ws[cell] = self.__iteration_date_string__(iteration_list[i])
            ws[cell].alignment = Alignment(horizontal='general', vertical='bottom', text_rotation=0, wrap_text=True,
                                           shrink_to_fit=False, indent=0)

            cell = f'{cell_letter}3'
            ws[cell] = iteration_list[i].get(PythonCustomFields.CalculatedCapacity)
            ws[cell].protection = Protection(locked=False)

            cell = f'{cell_letter}4'
            ws[cell] = f'={cell_letter}{self.totals_row}'

            cell = f'{cell_letter}5'
            ws[cell] = f'={cell_letter}4 > {cell_letter}3'

            ws.column_dimensions[cell_letter].width = self.config.get('Forecaster').get('VisualFormats').get(
                'Overall').get('CellWidth')

        self.logger.debug('end')

    def __write_items__(self, work_item_detail_list, iteration_list, ws):
        self.logger.debug('start')
        row = self.excel_work_item_section_start_row

        ordered_items = self.__order_work_items__(work_item_detail_list)
        num_iterations = len(iteration_list)

        actual_capacity_rows = []
        epic_row = 0
        for i in range(0, len(ordered_items)):
            item = ordered_items[i]
            if item.get(WorkItemFields.WorkItemType) not in [WorkItemTypes.Epic,
                                                             WorkItemTypes.Feature]:
                continue

            work_item_title_row = row
            self.__write_work_item_title_record__(item, row, ws)
            self.__format_work_item_title_record__(item, row, ws, num_iterations)

            if item.get(WorkItemFields.WorkItemType) == WorkItemTypes.Epic:
                epic_row = row
                row += 1
                continue

            row += 1

            self.__write_work_item_iteration_capacity_percentage_record__(iteration_list, row, ws)

            row += 1

            self.__write_work_item_desired_iteration_capacity_record__(iteration_list, row, ws)

            row += 1

            self.__write_work_item_actual_iteration_effort_record__(actual_capacity_rows, iteration_list, row, ws)

            row += 1

            self.__write_work_item_remaining_effort_record__(item, iteration_list, row, ws)

            self.__format_epic_conditional_color__(epic_row, work_item_title_row, num_iterations, ws)

            row += self.excel_num_rows_between_features

        self.__write_total_effort_allocated_record__(actual_capacity_rows, iteration_list, ws)

        self.logger.debug('end')

    def __calculate_totals_row__(self, num_epics, num_features):
        totals_row = self.excel_work_item_section_start_row
        totals_row += num_epics
        features_row_count = num_features * (4 + self.excel_num_rows_between_features) + 15
        totals_row += features_row_count

        return totals_row

    def __order_work_items__(self, work_item_detail_list):
        result = []

        ordered_full_list = sorted(work_item_detail_list, key=lambda x: WorkItemFields.BacklogPriority)
        for epic in [x for x in ordered_full_list if x.get(WorkItemFields.WorkItemType) == WorkItemTypes.Epic]:
            result.append(epic)
            for feature in [x for x in ordered_full_list if
                            x.get(WorkItemFields.WorkItemType) == WorkItemTypes.Feature and x.get(
                                WorkItemFields.Parent) == epic.get(WorkItemFields.Id)]:
                result.append(feature)
        return result

    def __write_total_effort_allocated_record__(self, actual_capacity_rows, iteration_list, ws):
        """sum the cells where actual effort was recorded"""
        cell = f'A{self.totals_row}'
        ws[cell] = 'Total Allocated'
        for x in range(0, len(iteration_list)):
            cell_letter = self.__number_to_excel_column_letter__(self.excel_num_columns_from_header + x)

            actual_effort_cell_list = ','.join(
                [f'{cell_letter}{x}' for x in actual_capacity_rows]) if len(actual_capacity_rows) > 0 else '0'
            totals_formula = f'=sum({actual_effort_cell_list})'

            cell = f'{cell_letter}{self.totals_row}'
            ws[cell] = totals_formula

    def __write_work_item_remaining_effort_record__(self, item, iteration_list, row, ws):
        """Compare sum of actual work up to this column against the original remaining work"""
        cell_letter = self.__number_to_excel_column_letter__(0)
        cell = f'{cell_letter}{row}'
        ws[cell] = 'Remaining Effort'

        cell_letter = self.__number_to_excel_column_letter__(1)
        cell = f'{cell_letter}{row}'
        ws[cell] = item.get(PythonCustomFields.RemainingWork)
        total_remaining_work_cell = cell

        for x, iteration in enumerate(iteration_list):
            cell_letter = self.__number_to_excel_column_letter__(self.excel_num_columns_from_header + x)
            starting_cell_letter = self.__number_to_excel_column_letter__(self.excel_num_columns_from_header)
            cell = f'{cell_letter}{row}'

            sum_of_actual_til_now_formula = f'SUM({starting_cell_letter}{row - 1}:{cell_letter}{row - 1})'
            remaining_formula = f'=ROUND(IF({sum_of_actual_til_now_formula} >= {total_remaining_work_cell}, 0, {total_remaining_work_cell}-{sum_of_actual_til_now_formula}),0)'
            ws[cell] = remaining_formula

        # ws.row_dimensions[row].hidden = True

    def __write_work_item_actual_iteration_effort_record__(self, actual_capacity_rows, iteration_list, row, ws):
        """Compare desired effort against what's available"""
        cell_letter = self.__number_to_excel_column_letter__(0)
        cell = f'{cell_letter}{row}'
        ws[cell] = 'Actual Iteration Effort'

        for x in range(0, len(iteration_list)):
            cell_letter = self.__number_to_excel_column_letter__(self.excel_num_columns_from_header + x)
            previous_cell_letter = self.__number_to_excel_column_letter__(
                self.excel_num_columns_from_header + x - 1)
            previous_actual_capacity_cell_list = ','.join(
                [f'{cell_letter}{x}' for x in actual_capacity_rows]) if len(actual_capacity_rows) > 0 else '0'

            iteration_capacity_less_actual_effort_above_formula = f'{cell_letter}{self.iteration_capacity_row}-SUM({previous_actual_capacity_cell_list})'
            desired_effort_formula = f'{cell_letter}{row - 1}'
            actual_iteration_effort_formula = f'=ROUND(IF({iteration_capacity_less_actual_effort_above_formula}-{desired_effort_formula}>=0, {desired_effort_formula}, {iteration_capacity_less_actual_effort_above_formula}),0)'

            cell = f'{cell_letter}{row}'
            ws[cell] = actual_iteration_effort_formula
            ws[cell].protection = Protection(locked=False)

        actual_capacity_rows.append(row)

    def __write_work_item_desired_iteration_capacity_record__(self, iteration_list, row, ws):
        """Compare desired capacity against remaining effort"""
        cell_letter = self.__number_to_excel_column_letter__(0)
        cell = f'{cell_letter}{row}'
        ws[cell] = 'Desired Iteration Effort'

        for x in range(0, len(iteration_list)):
            cell_letter = self.__number_to_excel_column_letter__(self.excel_num_columns_from_header + x)
            previous_cell_letter = self.__number_to_excel_column_letter__(
                self.excel_num_columns_from_header + x - 1)

            desired_iteration_capacity_formula = f'=ROUND(MIN({cell_letter}{row - 1}*{cell_letter}{self.iteration_capacity_row},{previous_cell_letter}{row + 2}, B{row + 2}),0)'
            cell = f'{cell_letter}{row}'
            ws[cell] = desired_iteration_capacity_formula

        # ws.row_dimensions[row].hidden = True

    def __write_work_item_iteration_capacity_percentage_record__(self, iteration_list, row, ws):
        """This should come from the Epic or Feature (or parent) and have a default.
        Right now using a default of 15%"""
        cell_letter = self.__number_to_excel_column_letter__(0)
        cell = f'{cell_letter}{row}'
        ws[cell] = '% of Iteration Capacity'
        percent_work_per_iteration = self.config.get('Forecaster').get('Iterations').get('Feature').get(
            'DefaultDesiredCapacity')
        for x in range(0, len(iteration_list)):
            cell_letter = self.__number_to_excel_column_letter__(self.excel_num_columns_from_header + x)
            cell = f'{cell_letter}{row}'
            ws[cell] = percent_work_per_iteration
            ws[cell].number_format = numbers.BUILTIN_FORMATS[10]  # 10: '0.00%',
            ws[cell].protection = Protection(locked=False)

    def __write_work_item_title_record__(self, item, row, ws):
        cell_letter = self.__number_to_excel_column_letter__(0)
        cell = f'{cell_letter}{row}'
        ws[cell] = item.get(WorkItemFields.WorkItemType)
        cell_letter = self.__number_to_excel_column_letter__(self.excel_num_columns_from_header - 1)
        cell = f'{cell_letter}{row}'
        ws[cell] = item.get(WorkItemFields.Title)

    def __format_work_item_title_record__(self, item, row, ws, num_iterations):
        work_item_type = item.get(WorkItemFields.WorkItemType)
        format = self.config.get('Forecaster').get('VisualFormats').get(work_item_type).get('Title')
        pattern_fill = PatternFill(**format.get('PatternFill'))
        font = Font(**format.get('Font'))
        for i in range(0, self.excel_num_columns_from_header):
            cell_letter = self.__number_to_excel_column_letter__(i)
            cell = f'{cell_letter}{row}'
            ws[cell].fill = pattern_fill
            ws[cell].font = font

        if work_item_type == WorkItemTypes.Epic:
            return

        for i in range(self.excel_num_columns_from_header, self.excel_num_columns_from_header + num_iterations):
            cell_letter = self.__number_to_excel_column_letter__(i)
            cell = f'{cell_letter}{row}'
            iteration_value_row = row + 3
            iteration_value_cell = f'{cell_letter}{iteration_value_row}'
            format = self.config.get('Forecaster').get('VisualFormats').get(work_item_type).get(
                'DoingWorkInIteration')
            pattern_fill = PatternFill(**format.get('PatternFill'))
            ws.conditional_formatting.add(f'{cell}:{cell}',
                                          FormulaRule(formula=[f'{iteration_value_cell}>0'], fill=pattern_fill))

    def __format_epic_conditional_color__(self, epic_row, feature_title_row, num_iterations, ws):
        formula_row_list = []
        next_feature_title_row = epic_row + 1
        while next_feature_title_row <= feature_title_row:
            formula_row_list.append(str(next_feature_title_row + 3))
            next_feature_title_row = next_feature_title_row + 4 + self.excel_num_rows_between_features

        for i in range(self.excel_num_columns_from_header, self.excel_num_columns_from_header + num_iterations):
            cell_letter = self.__number_to_excel_column_letter__(i)
            cell = f'{cell_letter}{epic_row}'
            formula_row_comma_delimited = ','.join([f'{cell_letter}{x}' for x in formula_row_list])
            conditional_formula = f'sum({formula_row_comma_delimited}) > 0'
            format = self.config.get('Forecaster').get('VisualFormats').get(WorkItemTypes.Epic).get(
                'DoingWorkInIteration')
            pattern_fill = PatternFill(**format.get('PatternFill'))
            ws.conditional_formatting.add(f'{cell}:{cell}',
                                          FormulaRule(formula=[conditional_formula], fill=pattern_fill))

    def __number_to_excel_column_letter__(self, number):
        """
        Convert 0-based index to its capital ascii character; to be used for Excel row identification.
        0 = A
        1 = B
        etc.
        :param number: 0-based index integer to be converted to ascii character
        :return: single letter (A-Z) or double-letter (ex: AB, DG, etc.)
        """
        string = ""
        number += 1
        while number > 0:
            number, remainder = divmod(number - 1, 26)
            string = chr(65 + remainder) + string
        return string

    def __iteration_date_string__(self, iteration):
        start_string = iteration.get(IterationFields.StartDate).strftime("%d %b")
        end_string = iteration.get(IterationFields.FinishDate).strftime("%d %b")
        result = f'{start_string} - {end_string}'
        return result
